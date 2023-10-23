from django.shortcuts import render, redirect, HttpResponse
from app01 import models
from app01.utils.pagination import Pagination


def depart_list(request):
    """部门列表"""

    # 获取所有部门列表
    # [object, object, object]
    queryset = models.Department.objects.all()

    page_object = Pagination(request, queryset, page_size=8)


    context= {
        "queryset":page_object.page_queryset,
        "page_string":page_object.html()
        }

    return render(request, 'depart_list.html', context)


def depart_add(request):
    """添加部门"""
    if request.method=="GET":
        return render(request, "depart_add.html")
    title = request.POST.get("title")
    # 保存到数据库 
    models.Department.objects.create(title=title)
    return redirect("/depart/list/") 


def depart_delete(request):
    """删除部门"""
    nid = request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")


def depart_edit(request, nid):
    """编辑部门"""
    if request.method=="GET":
        row_object = models.Department.objects.filter(id=nid).first()
        return render(request, "depart_edit.html", {"row_object":row_object})
    # 找到title
    title = request.POST.get("title")
    # 在数据库中更新title
    models.Department.objects.filter(id = nid).update(title=title)
    # 重定向
    return redirect("/depart/list/")


"""excel上传"""
def depart_multi(request):

    from openpyxl import load_workbook

    # 1. 获取用户上传的文件对象
    file_object = request.FILES.get('exc')

    # 2. 对象传递给 opnepyxl，由 openpyxl 读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 3，循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title = text).exists()
        if not exists:
            models.Department.objects.create(title = text)

    return redirect('/depart/list/')










